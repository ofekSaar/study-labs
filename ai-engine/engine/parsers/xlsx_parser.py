"""
Excel Spreadsheet Parser — Extracts cell data from XLSX files as Markdown tables.

Uses openpyxl for structured XML parsing.
No OCR needed — Excel stores cell values as structured data.
"""

import io
import logging
from typing import List

from openpyxl import load_workbook

from .base import BaseParser, ParseResult

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):

    @staticmethod
    def supported_extensions() -> List[str]:
        return ['.xlsx']

    def parse(self, file_bytes: bytes, filename: str) -> ParseResult:
        logger.info(f"XlsxParser: Processing '{filename}'...")
        text_parts = []
        sheet_names = []

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_names.append(sheet_name)
                
                rows_data = []
                for row in ws.iter_rows(values_only=True):
                    # Convert each cell to string, handle None
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    # Skip completely empty rows
                    if any(v.strip() for v in row_values):
                        rows_data.append(row_values)

                if not rows_data:
                    continue

                # Build Markdown table
                md_lines = []
                md_lines.append(f"## Sheet: {sheet_name}")
                md_lines.append("")

                # Header row
                header = rows_data[0]
                md_lines.append("| " + " | ".join(header) + " |")
                md_lines.append("| " + " | ".join(["---"] * len(header)) + " |")

                # Data rows
                for row_values in rows_data[1:]:
                    # Pad or trim to match header column count
                    padded = row_values[:len(header)]
                    while len(padded) < len(header):
                        padded.append("")
                    md_lines.append("| " + " | ".join(padded) + " |")

                text_parts.append("\n".join(md_lines))

            wb.close()

        except Exception as e:
            logger.error(f"XlsxParser Error for '{filename}': {e}")
            return ParseResult(text="", images=[], metadata={"error": str(e)})

        full_text = "\n\n".join(text_parts)
        logger.info(f"XlsxParser: Extracted {len(full_text)} chars from {len(sheet_names)} sheets in '{filename}'")

        return ParseResult(
            text=full_text,
            images=[],  # Excel doesn't typically have extractable images via openpyxl
            metadata={"sheet_count": len(sheet_names), "sheet_names": sheet_names}
        )
